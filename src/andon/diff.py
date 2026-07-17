"""Check-aware workbook diff: what changed between two versions, and does it matter?

`git diff` on an .xlsx is noise — a repacked zip of XML. The tools that do
compare spreadsheets show every changed cell flat, which drowns the one number
that moved 40% in a sea of reformatted dates. andon's diff is opinionated
instead: it classifies each change (a number moved, text changed, a formula was
edited, a cell appeared or vanished, a new error showed up), quantifies numeric
moves with a delta and a percent, and lets a tolerance hide the changes too
small to care about. A new `#REF!` in the later version is called out on its
own, because that is the change most likely to be a mistake.

This is not a check family — it does not run against a spec. It answers the
question a spec can't: "someone edited this workbook; what actually changed?"
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from andon.errors import SourceError
from andon.tolerance import Tolerance, parse_tolerance

ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

# Change kinds, ordered by how much a reader should care (most first).
KIND_ORDER = ["new_error", "numeric", "formula", "text", "appeared", "vanished"]


@dataclass
class CellChange:
    sheet: str
    coord: str
    kind: str
    before: Any
    after: Any
    delta: float | None = None
    pct: float | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "coord": self.coord,
            "kind": self.kind,
            "before": _jsonable(self.before),
            "after": _jsonable(self.after),
            "delta": self.delta,
            "pct": self.pct,
        }


@dataclass
class DiffReport:
    before: str
    after: str
    changes: list[CellChange] = field(default_factory=list)
    sheets_added: list[str] = field(default_factory=list)
    sheets_removed: list[str] = field(default_factory=list)
    cells_compared: int = 0
    tolerance: str | None = None

    def count(self, kind: str) -> int:
        return sum(1 for c in self.changes if c.kind == kind)

    @property
    def has_new_errors(self) -> bool:
        return self.count("new_error") > 0

    def exit_code(self, strict: bool = False) -> int:
        """0 = nothing meaningful changed. 1 = a new error appeared (the change
        most likely to be a bug), or, with --strict, any change at all. 2 =
        changes exist but none is a new error."""
        if self.has_new_errors:
            return 1
        if self.changes:
            return 1 if strict else 2
        if self.sheets_added or self.sheets_removed:
            return 1 if strict else 2
        return 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "before": self.before,
            "after": self.after,
            "tolerance": self.tolerance,
            "cells_compared": self.cells_compared,
            "sheets_added": self.sheets_added,
            "sheets_removed": self.sheets_removed,
            "counts": {k: self.count(k) for k in KIND_ORDER},
            "changes": [c.to_dict() for c in self.changes],
        }


def _jsonable(v: Any) -> Any:
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _is_error(v: Any) -> bool:
    return isinstance(v, str) and v in ERROR_VALUES


def _num(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def diff_workbooks(
    before: str | Path,
    after: str | Path,
    *,
    sheets: list[str] | None = None,
    tolerance: object = None,
) -> DiffReport:
    """Compare two workbooks cell by cell and classify what changed.

    Numeric changes within `tolerance` (an absolute number or a "0.5%" string)
    are treated as unchanged, so reformatting and rounding noise stays out of
    the report."""
    bpath, apath = Path(before), Path(after)
    for p in (bpath, apath):
        if not p.is_file():
            raise SourceError(f"Workbook not found: {p}")

    tol: Tolerance | None = None
    if tolerance is not None:
        tol = parse_tolerance(tolerance, Tolerance("abs", 0.0))

    bval = load_workbook(bpath, data_only=True, read_only=True)
    aval = load_workbook(apath, data_only=True, read_only=True)
    bfml = load_workbook(bpath, data_only=False, read_only=True)
    afml = load_workbook(apath, data_only=False, read_only=True)

    report = DiffReport(
        before=bpath.name,
        after=apath.name,
        tolerance=tol.describe() if tol else None,
    )

    b_sheets = set(bval.sheetnames)
    a_sheets = set(aval.sheetnames)
    common = [s for s in aval.sheetnames if s in b_sheets]  # after's order
    if sheets is not None:
        unknown = [s for s in sheets if s not in a_sheets and s not in b_sheets]
        if unknown:
            raise SourceError(f"Worksheet(s) not found in either workbook: {', '.join(unknown)}")
        common = [s for s in common if s in sheets]

    report.sheets_added = sorted(a_sheets - b_sheets)
    report.sheets_removed = sorted(b_sheets - a_sheets)

    for sheet in common:
        _diff_sheet(sheet, bval[sheet], aval[sheet], bfml[sheet], afml[sheet], tol, report)

    order = {k: i for i, k in enumerate(KIND_ORDER)}
    report.changes.sort(key=lambda c: (order.get(c.kind, 99), c.sheet, c.coord))
    return report


def _diff_sheet(name, bws, aws, bws_f, aws_f, tol, report) -> None:  # type: ignore[no-untyped-def]
    b_vals = _cells(bws)
    a_vals = _cells(aws)
    b_fmls = _cells(bws_f)
    a_fmls = _cells(aws_f)
    # Formula cells have no cached value (None in the data_only view), so the
    # address set must include the formula view too, or formula-only changes
    # would never be seen.
    coords = set(b_vals) | set(a_vals) | set(b_fmls) | set(a_fmls)
    report.cells_compared += len(coords)

    for coord in coords:
        bv = b_vals.get(coord)
        av = a_vals.get(coord)
        bf = b_fmls.get(coord)
        af = a_fmls.get(coord)

        # 1. A newly appeared error value is the headline change.
        if _is_error(av) and not _is_error(bv):
            report.changes.append(CellChange(name, coord, "new_error", bv, av))
            continue

        # 2. Formula cells are compared by formula, not by cached value (which
        #    openpyxl often can't read and which is derivative anyway — a
        #    changed input shows up on the input cell). formula==formula → no
        #    change even if a stale cached value differs.
        b_is_f = isinstance(bf, str) and bf.startswith("=")
        a_is_f = isinstance(af, str) and af.startswith("=")
        if b_is_f or a_is_f:
            if bf != af:
                report.changes.append(CellChange(name, coord, "formula", bf, af))
            continue

        # 3. Value cells.
        if bv is None and av is not None:
            report.changes.append(CellChange(name, coord, "appeared", None, av))
            continue
        if bv is not None and av is None:
            report.changes.append(CellChange(name, coord, "vanished", bv, None))
            continue
        if bv is None and av is None:
            continue

        bn, an = _num(bv), _num(av)
        if bn is not None and an is not None:
            if bn == an:
                continue
            delta = an - bn
            if tol is not None and tol.allows(an, bn):
                continue  # within tolerance: noise
            pct = round(100.0 * delta / bn, 2) if bn != 0 else None
            report.changes.append(
                CellChange(name, coord, "numeric", bv, av, round(delta, 6), pct)
            )
            continue

        if bv != av:
            report.changes.append(CellChange(name, coord, "text", bv, av))


def _cells(ws: Any) -> dict[str, Any]:
    """Map coord -> value for every non-empty cell in a read-only worksheet."""
    out: dict[str, Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                out[f"{get_column_letter(cell.column)}{cell.row}"] = cell.value
    return out
