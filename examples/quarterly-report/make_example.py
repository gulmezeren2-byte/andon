"""Stage the crime scene.

This script generates the example deterministically (seeded RNG), so the
numbers in the README are reproducible. It builds:

  data/orders.csv   about 560 orders for Q2 2026 — the source of truth
  report.xlsx       the "AI-drafted" quarterly report, with planted defects
                    of the kind I keep meeting in real reports:

  1. Summary!B3  order count taken from a stale snapshot (+12 cancelled)
  2. Summary!B4  net revenue typed over, 1.8% higher than the data
  3. Summary!B9:B12  region shares that sum to 101.2
  4. Summary!B18 a total row that predates the May data refresh
  5. Detail!D5   a constant typed over a formula column
  6. Detail!G2   a formula pointing at a deleted range (#REF!)
     Detail!F    freight numbers stored as text ("1.234,56")
     orders.csv  one refund posted as negative revenue (a REVIEW, not a FAIL)

Run me, then run andon:

    python make_example.py
    andon run andon.yaml
"""

from __future__ import annotations

import csv
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent
REGIONS = ["EU", "US", "APAC", "MEA"]
REGION_WEIGHTS = [0.42, 0.33, 0.18, 0.07]
Q2_START, Q2_DAYS = date(2026, 4, 1), 91


def build_orders(rng: random.Random) -> list[dict]:
    orders = []
    order_id = 10_000
    for day_offset in range(Q2_DAYS):
        day = Q2_START + timedelta(days=day_offset)
        for _ in range(rng.randint(4, 8)):  # every day trades: date_continuity passes
            order_id += 1
            region = rng.choices(REGIONS, weights=REGION_WEIGHTS)[0]
            revenue = round(rng.uniform(40, 900), 2)
            orders.append(
                {
                    "order_id": order_id,
                    "order_date": day.isoformat(),
                    "region": region,
                    "status": "cancelled" if rng.random() < 0.05 else "shipped",
                    "revenue": revenue,
                    "cost": round(revenue * rng.uniform(0.55, 0.8), 2),
                }
            )
    # One refund posted as negative revenue — plausible, worth a human look.
    orders[200]["status"] = "shipped"
    orders[200]["revenue"] = -45.0
    orders[200]["cost"] = 0.0
    return orders


def build_report(orders: list[dict]) -> None:
    shipped = [o for o in orders if o["status"] == "shipped"]
    total_revenue = sum(o["revenue"] for o in shipped)

    by_region = {r: 0.0 for r in REGIONS}
    for o in shipped:
        by_region[o["region"]] += o["revenue"]

    by_month = {"2026-04": 0.0, "2026-05": 0.0, "2026-06": 0.0}
    for o in shipped:
        by_month[o["order_date"][:7]] += o["revenue"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Quarterly Operations Report — Q2 2026"

    ws["A3"] = "Orders shipped"
    ws["B3"] = len(shipped) + 12  # defect 1: stale snapshot kept 12 cancelled orders

    ws["A4"] = "Net revenue"
    ws["B4"] = round(total_revenue * 1.018, 2)  # defect 2: typed-over, +1.8%

    ws["A5"] = "Average order value"
    ws["B5"] = round(total_revenue / len(shipped), 2)  # honest — this one passes

    ws["A8"], ws["B8"], ws["C8"] = "region", "share_pct", "revenue"
    for i, region in enumerate(REGIONS, start=9):
        share = round(100 * by_region[region] / total_revenue, 1)
        if region == "MEA":
            share = round(share + 1.3, 1)  # defect 3: shares now sum to 101.2
        ws[f"A{i}"] = region
        ws[f"B{i}"] = share
        ws[f"C{i}"] = round(by_region[region], 2)  # honest — group_sum passes

    ws["A14"], ws["B14"] = "month", "revenue"
    for i, (month, value) in enumerate(sorted(by_month.items()), start=15):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = round(value, 2)
    ws["A18"] = "Total"
    # defect 4: the total was computed before the May refresh added 8,432.10
    ws["B18"] = round(sum(by_month.values()) - 8_432.10, 2)

    det = wb.create_sheet("Detail")
    det["A1"], det["B1"], det["C1"], det["D1"], det["F1"] = (
        "segment", "revenue", "cost", "margin", "freight_manual",
    )
    seg_rows = [(r, m) for r in REGIONS[:2] for m in sorted(by_month)] + [("APAC", "2026-04"), ("APAC", "2026-05")]
    for i, (region, month) in enumerate(seg_rows, start=2):
        rev = round(by_region[region] / 3, 2)
        cost = round(rev * 0.68, 2)
        det[f"A{i}"] = f"{region} {month}"
        det[f"B{i}"] = rev
        det[f"C{i}"] = cost
        det[f"D{i}"] = f"=B{i}-C{i}"
    det["D5"] = 14_250.0  # defect 5: someone "fixed" a margin by typing over the formula
    det["G1"] = "check"
    det["G2"] = "=SUM(#REF!)"  # defect 6: the column this SUM pointed at was deleted
    det["F2"] = "1.234,56"  # freight keyed in as text, Turkish/European format
    det["F3"] = "897,40"

    wb.create_sheet("Scratch")["A1"] = "temp calculations, ignore"  # never referenced

    wb.save(HERE / "report.xlsx")


def main() -> None:
    rng = random.Random(42)
    orders = build_orders(rng)

    data_dir = HERE / "data"
    data_dir.mkdir(exist_ok=True)
    with open(data_dir / "orders.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(orders[0].keys()))
        writer.writeheader()
        writer.writerows(orders)

    build_report(orders)
    shipped = sum(1 for o in orders if o["status"] == "shipped")
    print(f"wrote data/orders.csv ({len(orders)} orders, {shipped} shipped) and report.xlsx")
    print("defects planted - now run: andon run andon.yaml")


if __name__ == "__main__":
    main()
