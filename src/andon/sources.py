"""Source loading: files in, DataFrames and cell values out.

Reference grammar (the string on the right-hand side of `sources:`):

    data/orders.csv                 CSV file
    data/orders.parquet             Parquet file (needs pyarrow installed)
    out/report.xlsx                 first worksheet of the workbook
    out/report.xlsx#Summary         a named worksheet
    out/report.xlsx#Summary!A1:D50  a rectangular range, first row = header

Inside a check, a *side* narrows a source further:

    { source: orders, where: "status != 'cancelled'" }   -> filtered table
    { source: report, cell: B4 }                         -> one cell value
    { source: report, range: "B10:B21" }                 -> numeric values
    { value: 1240 }                                      -> literal claim

Two behaviors here are deliberate and worth knowing:

* Header names are stripped of surrounding whitespace on load. Trailing
  spaces in Excel headers are invisible in the UI and have wasted enough
  human hours; this is the one normalization andon performs.
* Reading a `cell:` uses the value Excel last *calculated*, not the formula.
  If a workbook was produced by a library (openpyxl, xlsxwriter) and never
  recalculated, formula cells carry no cached value — andon refuses to guess
  and reports exactly that, because "probably fine" is not a verification.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string

from andon.errors import SourceError
from andon.result import SourceInfo

_REF_RE = re.compile(r"^(?P<path>[^#]+?)(?:#(?P<sheet>[^!]+?)(?:!(?P<range>[A-Za-z0-9:$]+))?)?$")
_CELL_RE = re.compile(r"^\$?[A-Za-z]{1,3}\$?[0-9]+$")


@dataclass(frozen=True)
class ParsedRef:
    path: Path
    sheet: str | None = None
    cell_range: str | None = None
    query: str | None = None  # a DuckDB SQL query, for `duckdb:` sources


def parse_ref(ref: str, base_dir: Path) -> ParsedRef:
    ref = ref.strip()
    # A `duckdb:` source is a SQL query, not a file. DuckDB can read CSV/parquet/
    # JSON and .duckdb files inside the query, so this turns "the report vs. a
    # warehouse query" into a first-class check. Relative paths in the SQL
    # resolve against the spec directory (path carries base_dir for that).
    if ref.lower().startswith("duckdb:"):
        query = ref[len("duckdb:") :].strip()
        if not query:
            raise SourceError("A `duckdb:` source needs a SQL query after the prefix.")
        return ParsedRef(path=base_dir, query=query)
    # '#' introduces a sheet only after an Excel path; '#' is a legal filename
    # character everywhere else (data/report#1.csv is a file, not a sheet ref).
    head = ref.split("#", 1)[0].strip()
    if "#" in ref and head.lower().endswith((".xlsx", ".xlsm")):
        m = _REF_RE.match(ref)
        if not m:
            raise SourceError(f"Cannot parse source reference: {ref!r}")
        path = Path(m.group("path").strip())
        sheet = m.group("sheet")
        cell_range = m.group("range")
    else:
        path = Path(ref)
        sheet = None
        cell_range = None
    if not path.is_absolute():
        path = base_dir / path
    return ParsedRef(path=path, sheet=sheet.strip() if sheet else None, cell_range=cell_range)


class SourceStore:
    """Loads and caches sources for one verification run. Read-only by construction:
    nothing in this class has a code path that writes to disk."""

    def __init__(self, base_dir: Path, sources: dict[str, str]) -> None:
        self.base_dir = base_dir
        self.refs = {alias: parse_ref(raw, base_dir) for alias, raw in sources.items()}
        self._raw = dict(sources)
        self._frames: dict[tuple[str, str | None], pd.DataFrame] = {}
        self._wb_values: dict[Path, Any] = {}
        self._wb_formulas: dict[Path, Any] = {}
        self._usage: dict[str, dict[str, Any]] = {}
        self.referenced_sheets: dict[Path, set[str]] = {}

    # -- public API used by checks ----------------------------------------

    def frame(self, alias: str, where: str | None = None) -> pd.DataFrame:
        ref = self._ref(alias)
        df = self._load_frame(alias, ref)
        if where:
            try:
                df = df.query(where, engine="python")
            except Exception as exc:
                raise SourceError(
                    f"Filter failed on source {alias!r}: {where!r} -> {exc}. "
                    f"Columns are: {', '.join(map(str, df.columns))}"
                ) from exc
        return df

    def cell(self, alias: str, cell: str) -> Any:
        ref = self._ref(alias)
        ws, sheet_name = self._worksheet(alias, ref, values=True)
        cell = cell.strip()
        if not _CELL_RE.match(cell):
            raise SourceError(f"{alias!r}: {cell!r} is not a cell address (like B4).")
        value = ws[cell.replace("$", "")].value
        if value is None and self._has_formula(ref.path, sheet_name, cell):
            raise SourceError(
                f"{alias}!{cell} holds a formula but the workbook carries no cached "
                f"value for it (it was written by a library and never recalculated). "
                f"Open and save the file in Excel/LibreOffice, or point the check at "
                f"the underlying data instead."
            )
        self._use(alias, ref, "excel")["cells"].append(f"{sheet_name}!{cell}")
        return value

    def values(self, alias: str, cell_range: str) -> tuple[list[float], int]:
        """Numeric values in a range, plus how many blank cells were skipped.
        Text in a numeric range is an error, not a zero."""
        ref = self._ref(alias)
        ws, sheet_name = self._worksheet(alias, ref, values=True)
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range.replace("$", ""))
        except ValueError as exc:
            raise SourceError(f"{alias!r}: bad range {cell_range!r}: {exc}") from exc

        numbers: list[float] = []
        blanks = 0
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for c in row:
                v = c.value
                if v is None:
                    blanks += 1
                elif isinstance(v, bool):
                    raise SourceError(f"{alias}!{c.coordinate} is a boolean in a numeric range.")
                elif isinstance(v, (int, float)):
                    numbers.append(float(v))
                else:
                    raise SourceError(
                        f"{alias}!{c.coordinate} is text ({v!r}) inside the numeric "
                        f"range {cell_range}. andon does not coerce text to numbers."
                    )
        self._use(alias, ref, "excel")["ranges"].append(f"{sheet_name}!{cell_range}")
        return numbers, blanks

    def workbook_path(self, alias: str) -> Path:
        ref = self._ref(alias)
        if ref.path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise SourceError(f"{alias!r} is not an Excel workbook.")
        return ref.path

    def workbooks(self, alias: str) -> tuple[Any, Any, Path]:
        """Both views of a workbook (cached values, formulas) plus its path.
        Loading through the store keeps the honesty block truthful: sheets the
        caller never touches show up under "Never read"."""
        path = self.workbook_path(alias)
        if not path.is_file():
            raise SourceError(f"Source file not found: {path}")
        return self._workbook(path, values=True), self._workbook(path, values=False), path

    def note_scan(self, alias: str, sheets: list[str]) -> None:
        """Record a whole-sheet scan (used by excel.integrity) so the report's
        "Read:" section reflects it."""
        ref = self._ref(alias)
        usage = self._use(alias, ref, "excel")
        usage["frames"].append("integrity scan: " + ", ".join(sheets))
        self.referenced_sheets.setdefault(ref.path, set()).update(sheets)

    def resolve_side(self, side: Any, *, context: str) -> tuple[str, Any]:
        """Resolve one side of a comparison.

        Returns ("frame", DataFrame) or ("scalar", value)."""
        if isinstance(side, (int, float)) and not isinstance(side, bool):
            return "scalar", float(side)
        if not isinstance(side, dict):
            raise SourceError(f"{context}: expected a mapping or a number, got {side!r}.")
        unknown = sorted(set(side) - {"source", "where", "cell", "value", "column"})
        if unknown:
            raise SourceError(
                f"{context}: unknown key(s) {', '.join(unknown)}. "
                f"A side takes: source, where, cell, value, column."
            )
        if "value" in side and len(side) > 1:
            raise SourceError(
                f"{context}: a literal `value` side takes no other keys "
                f"(got {', '.join(sorted(k for k in side if k != 'value'))})."
            )
        if "cell" in side and ("where" in side or "column" in side):
            raise SourceError(
                f"{context}: a `cell` side takes only `source` — "
                f"`where`/`column` would be silently meaningless."
            )
        if "value" in side:
            v = side["value"]
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise SourceError(f"{context}: `value` must be a number, got {v!r}.")
            return "scalar", float(v)
        alias = side.get("source")
        if not alias:
            raise SourceError(f"{context}: needs `source:` (or a literal `value:`).")
        if "cell" in side:
            v = self.cell(alias, side["cell"])
            if v is None:
                raise SourceError(f"{context}: {alias}!{side['cell']} is empty.")
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                raise SourceError(
                    f"{context}: {alias}!{side['cell']} is {v!r}, not a number."
                )
            return "scalar", float(v)
        return "frame", self.frame(alias, side.get("where"))

    # -- honesty-block support ----------------------------------------------

    def infos(self) -> list[SourceInfo]:
        """One line per alias, composed from everything that was actually read.
        Paths are shown relative to the spec so reports don't leak local
        directory layouts into logs, CI comments or screenshots."""
        out: list[SourceInfo] = []
        for alias, usage in self._usage.items():
            ref: ParsedRef = usage["ref"]
            if ref.query is not None:
                display = "(DuckDB query)"
            else:
                try:
                    display = ref.path.relative_to(self.base_dir).as_posix()
                except ValueError:
                    display = str(ref.path)
            parts: list[str] = list(dict.fromkeys(usage["frames"]))
            cells = list(dict.fromkeys(usage["cells"]))
            if cells:
                more = f" +{len(cells) - 8} more" if len(cells) > 8 else ""
                parts.append("cells " + ", ".join(cells[:8]) + more)
            ranges = list(dict.fromkeys(usage["ranges"]))
            if ranges:
                parts.append("ranges " + ", ".join(ranges[:8]))
            out.append(
                SourceInfo(
                    alias=alias,
                    path=display,
                    kind=usage["kind"],
                    detail="; ".join(parts),
                    rows=usage["rows"],
                )
            )
        return out

    def unreferenced_sheets(self) -> list[str]:
        """Worksheets that exist in referenced workbooks but were never read."""
        out: list[str] = []
        for path, used in sorted(self.referenced_sheets.items(), key=lambda kv: str(kv[0])):
            wb = self._wb_values.get(path)
            if wb is None:
                continue
            unused = [s for s in wb.sheetnames if s not in used]
            out.extend(f"{path.name}#{s}" for s in unused)
        return out

    # -- internals -----------------------------------------------------------

    def _ref(self, alias: str) -> ParsedRef:
        try:
            return self.refs[alias]
        except KeyError:
            known = ", ".join(sorted(self.refs)) or "(none)"
            raise SourceError(
                f"Unknown source alias {alias!r}. Declared sources: {known}."
            ) from None

    def _load_frame(self, alias: str, ref: ParsedRef) -> pd.DataFrame:
        key = (alias, ref.cell_range)
        if key in self._frames:
            return self._frames[key]

        if ref.query is not None:
            df = self._load_duckdb(ref)
            df.columns = [str(c).strip() for c in df.columns]
            self._frames[key] = df
            usage = self._use(alias, ref, "duckdb")
            usage["rows"] = len(df)
            q = ref.query if len(ref.query) <= 60 else ref.query[:57] + "..."
            usage["frames"].append(f"query: {q}")
            return df

        if not ref.path.is_file():
            raise SourceError(f"Source file not found: {ref.path}")

        suffix = ref.path.suffix.lower()
        if suffix == ".csv":
            try:
                df = pd.read_csv(ref.path)
            except (pd.errors.EmptyDataError, pd.errors.ParserError) as exc:
                raise SourceError(f"Cannot read {ref.path.name} as CSV: {exc}") from exc
            kind = "csv"
            detail = ""
        elif suffix == ".parquet":
            try:
                df = pd.read_parquet(ref.path)
            except ImportError as exc:
                raise SourceError(
                    "Reading parquet needs pyarrow: pip install 'andon[parquet]'"
                ) from exc
            kind = "parquet"
            detail = ""
        elif suffix in (".xlsx", ".xlsm"):
            df, detail = self._load_excel_frame(alias, ref)
            kind = "excel"
        else:
            raise SourceError(f"Unsupported source type {suffix!r} for {alias!r}.")

        df.columns = [str(c).strip() for c in df.columns]
        self._frames[key] = df
        usage = self._use(alias, ref, kind)
        usage["rows"] = len(df)
        if detail:
            usage["frames"].append(detail)
        return df

    def _load_duckdb(self, ref: ParsedRef) -> pd.DataFrame:
        """Run a `duckdb:` source's SQL and return the result as a DataFrame.
        The query runs with the spec directory as the working directory, so
        relative file paths inside the SQL (FROM 'data/orders.csv') resolve
        against the spec, not against wherever andon happened to be launched."""
        try:
            import duckdb
        except ImportError as exc:
            raise SourceError(
                "DuckDB sources need the optional dependency: "
                "pip install 'andon-verify[duckdb]'"
            ) from exc
        conn = duckdb.connect(":memory:")
        cwd = os.getcwd()
        try:
            os.chdir(ref.path)  # ref.path is the spec's base_dir for duckdb sources
            return conn.execute(str(ref.query)).df()
        except duckdb.Error as exc:
            raise SourceError(f"DuckDB query failed: {exc}") from exc
        finally:
            os.chdir(cwd)
            conn.close()

    def _load_excel_frame(self, alias: str, ref: ParsedRef) -> tuple[pd.DataFrame, str]:
        ws, sheet_name = self._worksheet(alias, ref, values=True)
        if ref.cell_range:
            min_col, min_row, max_col, max_row = range_boundaries(ref.cell_range.replace("$", ""))
            rows = list(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )
            if not rows:
                raise SourceError(f"{alias!r}: range {ref.cell_range} is empty.")
            header = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
            df = pd.DataFrame(rows[1:], columns=header)
            for col in df.columns:
                converted = pd.to_numeric(df[col], errors="coerce")
                if converted.notna().sum() == df[col].notna().sum():
                    df[col] = converted
            return df, f"{sheet_name}!{ref.cell_range}"
        df = pd.read_excel(ref.path, sheet_name=sheet_name)
        return df, sheet_name

    def _worksheet(self, alias: str, ref: ParsedRef, *, values: bool) -> tuple[Any, str]:
        if ref.path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise SourceError(
                f"{alias!r}: cell/range access needs an Excel source, got {ref.path.name}."
            )
        wb = self._workbook(ref.path, values=values)
        if ref.sheet:
            if ref.sheet not in wb.sheetnames:
                raise SourceError(
                    f"{alias!r}: worksheet {ref.sheet!r} not found in {ref.path.name}. "
                    f"Sheets: {', '.join(wb.sheetnames)}."
                )
            name = ref.sheet
        else:
            name = wb.sheetnames[0]
        self.referenced_sheets.setdefault(ref.path, set()).add(name)
        return wb[name], name

    def _workbook(self, path: Path, *, values: bool) -> Any:
        cache = self._wb_values if values else self._wb_formulas
        if path not in cache:
            if not path.is_file():
                raise SourceError(f"Source file not found: {path}")
            cache[path] = load_workbook(path, data_only=values, read_only=False)
        return cache[path]

    def _has_formula(self, path: Path, sheet: str, cell: str) -> bool:
        wb = self._workbook(path, values=False)
        col_row = coordinate_from_string(cell.replace("$", ""))
        v = wb[sheet][f"{col_row[0]}{col_row[1]}"].value
        return isinstance(v, str) and v.startswith("=")

    def _use(self, alias: str, ref: ParsedRef, kind: str) -> dict[str, Any]:
        return self._usage.setdefault(
            alias,
            {"ref": ref, "kind": kind, "rows": None, "frames": [], "cells": [], "ranges": []},
        )
