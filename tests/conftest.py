"""Shared fixtures: build real files, run real specs.

andon's unit of behavior is "spec in, verdict out", so most tests here go
through the same door a user does: write actual CSV/XLSX files to tmp_path,
run a spec dict against them, and assert on the resulting Report. No mocks —
a verifier that is only ever tested against fakes verifies nothing.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from andon.engine import run
from andon.result import Report
from andon.spec import Spec

ORDERS_CSV = """order_id,region,status,revenue,cost,order_date
1001,EU,shipped,120.0,80.0,2026-01-05
1002,EU,shipped,80.5,60.0,2026-01-06
1003,US,shipped,200.0,150.0,2026-01-07
1004,US,cancelled,50.0,40.0,2026-01-08
1005,EU,shipped,99.5,70.0,2026-01-09
1006,APAC,shipped,150.0,90.0,2026-01-10
"""
# shipped rows: 5 · shipped revenue: 650.0 · EU shipped revenue: 300.0


@pytest.fixture()
def orders_csv(tmp_path: Path) -> Path:
    p = tmp_path / "orders.csv"
    p.write_text(ORDERS_CSV, encoding="utf-8")
    return p


@pytest.fixture()
def workbook_factory(tmp_path: Path):
    """Build an .xlsx from {sheet: {cell: value}} mappings."""

    def build(name: str, sheets: dict[str, dict[str, Any]]) -> Path:
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, cells in sheets.items():
            ws = wb.create_sheet(sheet_name)
            for coord, value in cells.items():
                ws[coord] = value
        path = tmp_path / name
        wb.save(path)
        return path

    return build


@pytest.fixture()
def run_spec(tmp_path: Path):
    """Run a spec dict with tmp_path as the base directory."""

    def _run(data: dict[str, Any]) -> Report:
        spec = Spec.from_dict(data, base_dir=tmp_path, path="<test>")
        return run(spec)

    return _run


def one(report: Report):
    assert len(report.results) == 1, [r.summary for r in report.results]
    return report.results[0]
