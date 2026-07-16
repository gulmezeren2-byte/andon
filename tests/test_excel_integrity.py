from conftest import one

from andon.result import Status


def integrity_spec(path, **params) -> dict:
    return {
        "version": 1,
        "sources": {"wb": path.name},
        "checks": [{"excel.integrity": {"source": "wb", **params}}],
    }


def test_clean_workbook_passes(workbook_factory, run_spec) -> None:
    wb = workbook_factory("clean.xlsx", {"S": {"A1": "x", "A2": 1, "A3": 2}})
    report = run_spec(integrity_spec(wb))
    assert one(report).status is Status.PASS


def test_error_value_fails(workbook_factory, run_spec) -> None:
    wb = workbook_factory("broken.xlsx", {"S": {"A1": 1, "B2": "#REF!"}})
    report = run_spec(integrity_spec(wb))
    r = one(report)
    assert r.status is Status.FAIL
    assert "S!B2" in r.evidence["error_cells"][0]


def test_ref_error_inside_formula_fails(workbook_factory, run_spec) -> None:
    wb = workbook_factory("brokenf.xlsx", {"S": {"A1": 1, "B2": "=SUM(#REF!)"}})
    report = run_spec(integrity_spec(wb))
    r = one(report)
    assert r.status is Status.FAIL
    assert r.evidence["counts"]["error_cells"] == 1


def test_hardcoded_value_in_formula_column_reviewed(workbook_factory, run_spec) -> None:
    cells = {f"C{i}": f"=A{i}*2" for i in range(2, 8)}  # six formulas
    cells["C9"] = 999.0  # one typed-over constant
    cells["A1"] = "x"
    wb = workbook_factory("hard.xlsx", {"S": cells})
    report = run_spec(integrity_spec(wb))
    r = one(report)
    assert r.status is Status.REVIEW
    assert r.evidence["hardcoded_in_formula_columns"] == ["S!C9 = 999"]


def test_number_stored_as_text_reviewed(workbook_factory, run_spec) -> None:
    wb = workbook_factory("text.xlsx", {"S": {"A1": 10, "A2": "1.234,56", "A3": "42"}})
    report = run_spec(integrity_spec(wb))
    r = one(report)
    assert r.status is Status.REVIEW
    assert r.evidence["counts"]["numeric_text"] == 2


def test_merged_cells_are_informational_only(workbook_factory, run_spec) -> None:
    import openpyxl

    wb_path = workbook_factory("merged.xlsx", {"S": {"A1": "title", "A3": 1, "B3": 2}})
    wb = openpyxl.load_workbook(wb_path)
    wb["S"].merge_cells("A1:B1")
    wb.save(wb_path)

    report = run_spec(integrity_spec(wb_path))
    r = one(report)
    assert r.status is Status.PASS  # merged alone must not flag
    assert r.evidence["counts"]["merged"] == 1


def test_hidden_row_reviewed(workbook_factory, run_spec) -> None:
    import openpyxl

    wb_path = workbook_factory("hidden.xlsx", {"S": {"A1": 1, "A2": 2, "A3": 3}})
    wb = openpyxl.load_workbook(wb_path)
    wb["S"].row_dimensions[2].hidden = True
    wb.save(wb_path)

    report = run_spec(integrity_spec(wb_path))
    r = one(report)
    assert r.status is Status.REVIEW
    assert "S!row 2" in r.evidence["hidden_rows_cols"]


def test_unknown_sheet_is_an_error(workbook_factory, run_spec) -> None:
    wb = workbook_factory("s.xlsx", {"S": {"A1": 1}})
    report = run_spec(integrity_spec(wb, sheets=["Nope"]))
    assert one(report).status is Status.ERROR
