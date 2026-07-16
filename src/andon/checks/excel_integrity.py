"""Workbook integrity scan: is the Excel file mechanically sound?

Excel is where analysis goes to get edited by hand, and hand edits leave
scars. This check reads a workbook twice — once for formulas, once for the
values Excel last calculated — and looks for the scars:

* error cells (#REF!, #DIV/0!, ...) and formulas pointing at deleted ranges
  — deterministic, these FAIL;
* numeric constants sitting in a column that is otherwise formulas (someone
  typed over a formula to "fix" a number) — heuristic, REVIEW;
* numbers stored as text, including the Turkish/European "1.234,56" flavor
  that silently drops out of every SUM — heuristic, REVIEW;
* hidden rows/columns inside the data, and links into other workbooks —
  heuristic, REVIEW.

Merged cells and volatile formulas (NOW, TODAY, RAND, INDIRECT) are recorded
in the evidence but do not change the verdict: in real reports they are too
common to be a signal on their own.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string

from andon.checks import Outcome, register, reject_unknown, require
from andon.errors import SourceError, SpecError
from andon.result import Status
from andon.sources import SourceStore

ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
_EXTERNAL_REF = re.compile(r"\[[^\]]+\.xls[xm]?\]", re.IGNORECASE)
_VOLATILE = re.compile(r"\b(NOW|TODAY|RAND|RANDBETWEEN|INDIRECT|OFFSET)\s*\(", re.IGNORECASE)
_TR_NUMBER_TEXT = re.compile(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$|^-?\d+,\d+$")
_CAP = 20


def _looks_numeric_text(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return bool(_TR_NUMBER_TEXT.match(text))


@register("excel.integrity")
def integrity(store: SourceStore, params: dict[str, Any]) -> Outcome:
    kind = "excel.integrity"
    reject_unknown(params, ("source", "sheets"), kind)
    alias = require(params, "source", kind)
    only_sheets = params.get("sheets")
    if only_sheets is not None and (
        not isinstance(only_sheets, list) or not all(isinstance(s, str) for s in only_sheets)
    ):
        raise SpecError(f"{kind}: `sheets` must be a list of worksheet names.")

    wb_values, wb_formulas, path = store.workbooks(alias)

    sheet_names = only_sheets or wb_values.sheetnames
    unknown = [s for s in sheet_names if s not in wb_values.sheetnames]
    if unknown:
        raise SourceError(
            f"{kind}: worksheet(s) not found: {', '.join(unknown)}. "
            f"Sheets: {', '.join(wb_values.sheetnames)}."
        )
    store.note_scan(alias, list(sheet_names))

    error_cells: list[str] = []
    hardcoded: list[str] = []
    numeric_text: list[str] = []
    hidden: list[str] = []
    external: list[str] = []
    merged: list[str] = []
    volatile: list[str] = []
    counts = dict.fromkeys(
        ("error_cells", "hardcoded", "numeric_text", "hidden", "external", "merged", "volatile"),
        0,
    )

    def note(bucket: list[str], key: str, item: str) -> None:
        counts[key] += 1
        if len(bucket) < _CAP:
            bucket.append(item)

    for name in sheet_names:
        ws_v = wb_values[name]
        ws_f = wb_formulas[name]

        # Column profile from the formula view: which columns are formula-driven?
        col_formulas: dict[int, int] = {}
        col_constants: dict[int, list[Any]] = {}
        for row in ws_f.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    col_formulas[cell.column] = col_formulas.get(cell.column, 0) + 1
                    if _EXTERNAL_REF.search(v):
                        note(external, "external", f"{name}!{cell.coordinate}")
                    if "#REF!" in v:
                        note(error_cells, "error_cells", f"{name}!{cell.coordinate} ({v[:40]})")
                    if _VOLATILE.search(v):
                        note(volatile, "volatile", f"{name}!{cell.coordinate}")
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    col_constants.setdefault(cell.column, []).append(cell)

        for col, n_formula in col_formulas.items():
            constants = col_constants.get(col, [])
            total = n_formula + len(constants)
            if n_formula >= 5 and total and n_formula / total >= 0.6:
                for cell in constants:
                    note(hardcoded, "hardcoded", f"{name}!{cell.coordinate} = {cell.value}")

        # Value view: calculation errors and numbers stored as text.
        for row in ws_v.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    if v in ERROR_VALUES:
                        note(error_cells, "error_cells", f"{name}!{cell.coordinate} ({v})")
                    elif _looks_numeric_text(v):
                        note(numeric_text, "numeric_text", f"{name}!{cell.coordinate} = {v!r}")

        # Hidden rows/columns inside the used range.
        for idx, dim in ws_v.row_dimensions.items():
            if dim.hidden and idx <= (ws_v.max_row or 0):
                note(hidden, "hidden", f"{name}!row {idx}")
        for letter, dim in ws_v.column_dimensions.items():
            if dim.hidden and column_index_from_string(letter) <= (ws_v.max_column or 0):
                note(hidden, "hidden", f"{name}!col {letter}")

        for rng in list(ws_v.merged_cells.ranges):
            note(merged, "merged", f"{name}!{rng}")

    evidence: dict[str, Any] = {
        "workbook": path.name,
        "sheets_scanned": sheet_names,
        "counts": counts,
        "error_cells": error_cells,
        "hardcoded_in_formula_columns": hardcoded,
        "numbers_stored_as_text": numeric_text,
        "hidden_rows_cols": hidden,
        "external_references": external,
        "merged_ranges": merged,
        "volatile_formulas": volatile,
        "note": (
            "error cells are deterministic failures; hardcoded/numeric-text/hidden/"
            "external are heuristics (review); merged/volatile are informational only"
        ),
    }

    if counts["error_cells"]:
        return Outcome(
            Status.FAIL,
            f"{counts['error_cells']} error cell(s) in {path.name} "
            f"(first: {error_cells[0]})",
            evidence,
        )

    review_total = (
        counts["hardcoded"] + counts["numeric_text"] + counts["hidden"] + counts["external"]
    )
    if review_total:
        parts = []
        if counts["hardcoded"]:
            parts.append(f"{counts['hardcoded']} hardcoded value(s) in formula columns")
        if counts["numeric_text"]:
            parts.append(f"{counts['numeric_text']} number(s) stored as text")
        if counts["hidden"]:
            parts.append(f"{counts['hidden']} hidden row(s)/col(s)")
        if counts["external"]:
            parts.append(f"{counts['external']} external reference(s)")
        return Outcome(Status.REVIEW, "; ".join(parts), evidence)

    return Outcome(
        Status.PASS,
        f"{path.name}: no error cells, no hardcode/text-number/hidden findings "
        f"in {len(sheet_names)} sheet(s)",
        evidence,
    )
